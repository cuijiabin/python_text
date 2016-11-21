# coding=gbk
import os
import json
import requests
import datetime
from pyquery import PyQuery as pq
from openpyxl import Workbook
from openpyxl import load_workbook


#��ȡ����ID�б�
def get_cityId_list(url):
    city_list = []
    html = pq(url= url)
    for areaId in html.find('#filterCollapse').find('div[class="has-more workcity"]').eq(0).find('div[class="more more-positions"]').find("a[data-lg-tj-cid='idnull']"):
        aId = pq(areaId).attr('href').replace('http://www.lagou.com/gongsi/', '').replace('-0-0#filterBox', '')
        if(aId=='0'):
            continue
        city_list.append(aId)
    return city_list

#��ȡ���������б�
def get_city_name_list(u):
    city_name_list = []
    url = 'http://www.lagou.com/gongsi/'
    html = pq(url=url)
    for areaId in html.find('#filterCollapse').find('div[class="has-more workcity"]').eq(0).find('div[class="more more-positions"]').find("a[data-lg-tj-cid='idnull']"):
        area_name=pq(areaId).html()
        if area_name=="ȫ��":
            continue
        city_name_list.append(area_name)
    return city_name_list

#��ȡ������һ���ж���ҳ
def get_city_page(areaId,page_num):
    try:
        param = {'first': 'false', 'pn': page_num, 'sortField': '0', 'havemark': '0'} #���ʲ���
        r = requests.post('http://www.lagou.com/gongsi/'+areaId+'-0-0.json',params=param ) #requsets����
        page_num += 1
        if(len(r.json()['result'])/16==1):
            return get_city_page(areaId,page_num)
        else:
            return page_num
    except:
        return page_num-1

#���ݳ���ID��ȡ���й�˾��Ϣ
def get_company_list(areaId):
    company_list = []
    city_page_total=get_city_page(areaId,1)
    for pageIndex in range(1,city_page_total):
        print('������ȡ��'+str(pageIndex)+'ҳ')
        json_url = 'http://www.lagou.com/gongsi/'+areaId+'-0-0.json'
        param = {'first': 'false', 'pn': str(pageIndex), 'sortField': '0', 'havemark': '0'} #���ʲ���
        r = requests.post(json_url,params=param ) #requsets����
        msg = json.loads(r.text)
        try:
            for company in msg['result']:
               company_list.append([company['city'],company['cityScore'],company['companyFeatures'],company['companyId'],company['companyLabels'],company['companyLogo'],company['companyName'],str(company['companyPositions']),company['companyShortName'],company['countryScore'],company['createTime'],company['finaceStage'],company['industryField'],company['interviewRemarkNum'],company['otherLabels'], company['positionNum'],company['processRate'],str(datetime.datetime.now())])
        except:
            print('��ȡ���Ϊ'+str(areaId)+'����ʱ��'+str(pageIndex)+'ҳ�����˴���,����ʱ���󷵻�����Ϊ��'+str(msg))
            continue
    return company_list

#д��Excel�ļ�����
def write_file(fileName):
    # list = []
    # wb = Workbook()
    # ws = wb.active
    url = 'http://www.lagou.com/gongsi/'
    area_name_list = get_city_name_list(url)
    # for area_name in area_name_list:
    #     wb.create_sheet(title = area_name)
    #     file_name = fileName+'.xlsx'
    #     wb.save(file_name)
    areaId_list = get_cityId_list(url)
    for areaId in areaId_list:
        company_list = get_company_list(areaId)
        print('������ȡ----->****'+company_list[0][0]+'****��˾�б�')
        # wb1 = load_workbook(file_name)
        # ws = wb1.get_sheet_by_name(company_list[0][0])
        # ws.append(['��������','���е÷�','��˾����','��˾ID','��˾��ǩ','��˾Logo','��չ�׶�','��ҵ����','��ҵλ��','��ҵ���','ע��ʱ��','����״��','��ҵ','����ְλ','������ǩ','����������'])
        print(['��������','���е÷�','��˾����','��˾ID','��˾��ǩ','��˾Logo','��չ�׶�','��ҵ����','��ҵλ��','��ҵ���','ע��ʱ��','����״��','��ҵ','����ְλ','������ǩ','����������'])
    for company in company_list:
        # ws.append([company[0],str(company[1]),company[2],str(company[3]),company[4],company[5],company[6],company[7],company[8],company[9],company[10],company[11],company[12],company[13],company[14],company[15]])
        print([company[0],str(company[1]),company[2],str(company[3]),company[4],company[5],company[6],company[7],company[8],company[9],company[10],company[11],company[12],company[13],company[14],company[15]])
        # wb1.save(file_name)


# file_name =  input('�������ļ�����')
print(str(datetime.datetime.now()))
write_file("")
print(str(datetime.datetime.now()))
